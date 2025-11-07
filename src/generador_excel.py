"""
Módulo para generar el archivo Excel con formato
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class GeneradorExcel:
    def __init__(self, config):
        self.config = config
        self.estilos = self._definir_estilos()
    
    def _definir_estilos(self):
        """Define los estilos para el Excel"""
        return {
            'titulo': Font(bold=True, size=14),
            'encabezado': Font(bold=True, size=10),
            'normal': Font(size=9),
            'total': Font(bold=True, size=9),
            'centrado': Alignment(horizontal='center', vertical='center'),
            'borde_fino': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def crear_reporte_completo(self, datos):
        """Crea el reporte Excel completo"""
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Crear hojas
        self._crear_hoja_principal(wb, datos)
        self._crear_hojas_resumen(wb, datos)
        
        # Ajustar formato
        self._ajustar_formato_general(wb)
        
        # Guardar archivo
        archivo_salida = self._generar_nombre_archivo()
        wb.save(archivo_salida)
        
        return archivo_salida
    
    def _crear_hoja_principal(self, wb, datos):
        """Crea la hoja 'Pregrado CSFB'"""
        ws = wb.create_sheet("Pregrado CSFB")
        
        # Agregar títulos y encabezados
        self._agregar_titulos(ws)
        self._agregar_encabezados(ws)
        self._agregar_datos_carreras(ws, datos)
        self._agregar_totales(ws, datos)
    
    def _crear_hojas_resumen(self, wb, datos):
        """Crea las hojas de resumen"""
        hojas_resumen = [
            ("EVALUADOS", "evaluados"),
            ("ADMITIDOS", "admitidos"),
            ("MATRICULADOS", "matriculados"),
            ("ASIGNADOS", "asignados")
        ]
        
        for nombre_hoja, tipo in hojas_resumen:
            ws = wb.create_sheet(nombre_hoja)
            self._crear_hoja_resumen(ws, datos, tipo)
    
    def _generar_nombre_archivo(self):
        """Genera el nombre del archivo de salida"""
        fecha = datetime.now().strftime('%d_%m_%Y')
        nombre_base = f"REPORTE VRAC CAMPUS CENTRAL {fecha}.xlsx"
        return os.path.join(self.config.ruta_salida, nombre_base)
    
    # ... (métodos específicos de cada hoja)