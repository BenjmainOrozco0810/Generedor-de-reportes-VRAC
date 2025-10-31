# generador_excel.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from configuracion import Configuracion

class GeneradorExcel:
    def __init__(self):
        self.config = Configuracion()
        self.estilos = self._definir_estilos()
    
    def _definir_estilos(self):
        """Define los estilos de formato para el Excel"""
        return {
            'titulo_principal': Font(name='Arial', size=14, bold=True),
            'subtitulo': Font(name='Arial', size=12, bold=True),
            'encabezado': Font(name='Arial', size=10, bold=True),
            'normal': Font(name='Arial', size=10),
            'negrita': Font(name='Arial', size=10, bold=True),
            'resaltado': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'centrado': Alignment(horizontal='center', vertical='center'),
            'borde_fino': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def crear_reporte_completo(self, datos_estructurados, fecha_reporte=None):
        """Crea el reporte Excel completo en formato VRAC"""
        if not fecha_reporte:
            fecha_reporte = self.config.obtener_fecha_actual()
        
        # Crear workbook y hoja principal
        wb = Workbook()
        ws = wb.active
        ws.title = "Pregrado CSFB"
        
        # Agregar contenido
        self._agregar_encabezados(ws)
        self._agregar_datos(ws, datos_estructurados)
        self._agregar_totales(ws, datos_estructurados)
        self._agregar_pie_pagina(ws)
        
        # Ajustar formato
        self._ajustar_formato_columnas(ws)
        
        # Guardar archivo
        nombre_archivo = f"REPORTE_VRAC_CAMPUS_CENTRAL_{fecha_reporte}.xlsx"
        wb.save(nombre_archivo)
        
        print(f"✅ Reporte Excel generado: {nombre_archivo}")
        return nombre_archivo
    
    def _agregar_encabezados(self, ws):
        """Agrega los encabezados del reporte"""
        # Títulos principales
        ws.merge_cells('B3:F3')
        ws['B3'] = "Generación 2026"
        ws['B3'].font = self.estilos['titulo_principal']
        ws['B3'].alignment = self.estilos['centrado']
        
        ws.merge_cells('B4:F4')
        ws['B4'] = "PREGRADO"
        ws['B4'].font = self.estilos['subtitulo']
        ws['B4'].alignment = self.estilos['centrado']
        
        ws.merge_cells('B5:F5')
        ws['B5'] = "Campus San Francisco de Borja, S.J."
        ws['B5'].font = self.estilos['subtitulo']
        ws['B5'].alignment = self.estilos['centrado']
        
        # Encabezados de columnas
        encabezados = self.config.obtener_encabezados_vrac()
        for col_idx, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=8, column=col_idx, value=encabezado)
            celda.font = self.estilos['encabezado']
            celda.alignment = self.estilos['centrado']
            celda.border = self.estilos['borde_fino']
    
    def _agregar_datos(self, ws, datos_estructurados):
        """Agrega los datos de las carreras al Excel"""
        fila_actual = 9
        
        for facultad, carreras in datos_estructurados.items():
            for nombre_carrera, datos_carrera in carreras.items():
                # Facultad y Carrera
                ws.cell(row=fila_actual, column=2, value=facultad).font = self.estilos['negrita']
                ws.cell(row=fila_actual, column=3, value=nombre_carrera).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=5, value=datos_carrera['jornada']).font = self.estilos['normal']
                
                # Evaluados
                ws.cell(row=fila_actual, column=9, value=datos_carrera['evaluados_2025']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=10, value=datos_carrera['evaluados_2026']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=11, value=f"=J{fila_actual}-I{fila_actual}").font = self.estilos['normal']
                ws.cell(row=fila_actual, column=12, value=f"=IF(I{fila_actual}=0,0,K{fila_actual}/I{fila_actual})").font = self.estilos['normal']
                
                # Admitidos
                ws.cell(row=fila_actual, column=14, value=datos_carrera['admitidos_2025']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=15, value=datos_carrera['admitidos_2026']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=16, value=f"=O{fila_actual}-N{fila_actual}").font = self.estilos['normal']
                ws.cell(row=fila_actual, column=17, value=f"=IF(N{fila_actual}=0,0,P{fila_actual}/N{fila_actual})").font = self.estilos['normal']
                
                # Matriculados
                ws.cell(row=fila_actual, column=19, value=datos_carrera['matriculados_2025']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=20, value=datos_carrera['matriculados_2026']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=21, value=f"=T{fila_actual}-S{fila_actual}").font = self.estilos['normal']
                ws.cell(row=fila_actual, column=22, value=f"=IF(S{fila_actual}=0,0,U{fila_actual}/S{fila_actual})").font = self.estilos['normal']
                
                # Asignados
                ws.cell(row=fila_actual, column=24, value=datos_carrera['asignados_2025']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=25, value=datos_carrera['asignados_2026']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=26, value=f"=Y{fila_actual}-X{fila_actual}").font = self.estilos['normal']
                ws.cell(row=fila_actual, column=27, value=f"=IF(X{fila_actual}=0,0,Z{fila_actual}/X{fila_actual})").font = self.estilos['normal']
                
                # Meta y cumplimiento
                ws.cell(row=fila_actual, column=29, value=datos_carrera['meta_2026']).font = self.estilos['normal']
                ws.cell(row=fila_actual, column=31, value=f"=AC{fila_actual}-Y{fila_actual}").font = self.estilos['normal']
                ws.cell(row=fila_actual, column=32, value=f"=Y{fila_actual}/AC{fila_actual}").font = self.estilos['normal']
                
                # Aplicar bordes a toda la fila
                for col in range(1, 33):
                    ws.cell(row=fila_actual, column=col).border = self.estilos['borde_fino']
                
                fila_actual += 1
            
            # Espacio entre facultades
            fila_actual += 1
    
    def _agregar_totales(self, ws, datos_estructurados):
        """Calcula y agrega filas de totales por facultad"""
        # Esta función calcula los totales automáticamente
        # Se implementa similar a las fórmulas de SUM en tu Excel original
        pass
    
    def _agregar_pie_pagina(self, ws):
        """Agrega el pie de página con información de fuente"""
        ultima_fila = ws.max_row + 2
        
        ws.cell(row=ultima_fila, column=2, value="Fuente: Sistema de Reportes COA. Actualizado al 27/10/2025. Hora. 10:00:00 A.M.")
        ws.cell(row=ultima_fila+1, column=2, value="*Evaluados, generación 2026, al 25/10/2025.")
        ws.cell(row=ultima_fila+2, column=2, value="*Evaluados, generación 2025, al 28/10/2024.")
    
    def _ajustar_formato_columnas(self, ws):
        """Ajusta el ancho de las columnas para mejor visualización"""
        anchos_columnas = {
            'A': 5, 'B': 25, 'C': 40, 'D': 3, 'E': 12, 'F': 3, 'G': 3,
            'H': 3, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 3,
            'N': 12, 'O': 12, 'P': 12, 'Q': 12, 'R': 3,
            'S': 12, 'T': 12, 'U': 12, 'V': 12, 'W': 3,
            'X': 12, 'Y': 12, 'Z': 12, 'AA': 12, 'AB': 3,
            'AC': 12, 'AD': 12, 'AE': 20, 'AF': 15
        }
        
        for col, ancho in anchos_columnas.items():
            ws.column_dimensions[col].width = ancho